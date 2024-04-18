from database import conectar
from openpyxl import load_workbook
from propriedade import Propriedade

conn = conectar()
cursor = conn.cursor()

caminhoPlanilha = 'C:/Users/Isaías/Downloads/careiro-coordenadas.xlsx';

def consulta_select_propriedade_ss(filename):
    propriedades = []

    try:
        workbook = load_workbook(filename=filename)
        sheet = workbook.active

        for row in sheet.iter_rows(min_row=2, values_only=True):
            codigo, latitude, longitude = row
            propriedade = Propriedade(
                str(codigo), 
                formatar_coordenada(str(latitude)), 
                formatar_coordenada(str(longitude))
            )
            propriedades.append(propriedade)
        
        return propriedades

    except Exception as e:
        print("Algo de errado com os dados da planilha aconteceu", e)

def consulta_select_propriedade_db(codigo):
    try:
        query = f"select p.id_inscricaoestadual, p.nu_codigoanimal, i.vl_latitude, i.vl_longitude from agrocomum.propriedade p join agrocomum.inscricaoestadual i on i.id_inscricaoestadual = p.id_inscricaoestadual where p.nu_codigoanimal = '{codigo}'"
        cursor.execute(query)
        
        resultado = cursor.fetchone()
        
        return resultado
    except Exception as e:
        print("Erro ao executar consulta:", e)

def consulta_select_inscricaoestadual_db(id):
    try:
        query = f"select * from agrocomum.inscricaoestadual i where i.id_inscricaoestadual = '{id}'"
        cursor.execute(query)
        
        resultado = cursor.fetchone()
        
        return resultado
    except Exception as e:
        print("Erro ao executar consulta:", e)

def formatar_coordenada(coordenada):
    coordenada = coordenada.replace('-', '')
    coordenada = coordenada.replace(' ', '')
    coordenada = coordenada.replace("'", '')
    coordenada = coordenada.replace(",", '')
    coordenada = coordenada.replace("º", '')
    coordenada = coordenada.replace("°", '')
    
    return coordenada[0:7]

def atualizar_coordenadas_no_banco():
    propriedadesSS = consulta_select_propriedade_ss(caminhoPlanilha)
    count = 0;
    
    for item in propriedadesSS:
        try:
            print("Propriedade atual: ", item.codigo)

            propriedade = consulta_select_propriedade_db(item.codigo)
            update_inscricaoestadual_db(item.latitude, item.longitude, propriedade[0])
            
            print("Propriedade atualizada!")
            count = count + 1
        except Exception as e:
            print("Erro ao executar consulta:", e)
    
    print("|---")
    print("propriedades atualizadas: ", count)

def update_inscricaoestadual_db(latitude, longitude, id):
    try:
        query = f"UPDATE agrocomum.inscricaoestadual SET vl_latitude = '{latitude}', vl_longitude = '{longitude}' WHERE id_inscricaoestadual = '{id}'"
        cursor.execute(query)
        conn.commit()

        print("Coordenadas atualizadas no banco de dados com sucesso!")
    except Exception as e:
        conn.rollback()
        print("Erro ao atualizar coordenadas no banco de dados:", e)

atualizar_coordenadas_no_banco()

cursor.close()
conn.close()